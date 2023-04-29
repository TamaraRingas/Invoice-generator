import openpyxl
import datetime
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

#pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
pdfmetrics.registerFont(TTFont('VeraBd', 'VeraBd.ttf'))

#import excel file
wb = openpyxl.load_workbook('Invoices.xlsx')
sheet = wb['Sheet1']

page_height = 1000
page_width = 1000
margin = 50

def create_invoice():
    for i in range(2,7):
        invoice_number = int(sheet.cell(row = i, column = 1).value)
        parent = sheet.cell(row = i, column = 2).value
        lessons = int(sheet.cell(row = i, column = 3).value)
        price_per_lesson = int(sheet.cell(row = i, column = 4).value)
        total = lessons*price_per_lesson
        invoice_date = sheet.cell(row = i, column = 6).value

        doc = SimpleDocTemplate(str(parent) + str(invoice_date) + ".pdf", pagesize = letter)

        data = [ ['Kate de Gruchy Invoice for Piano Lessons', invoice_date.strftime('%m/%d/%Y')],
                 ['Invoice number', invoice_number],
                 ['Issued to', parent],
                 ['Lessons', lessons],
                 ['Price per lesson', price_per_lesson],
                 ['Total Due', total]      
        ]

        table = Table(data)
        elements = []
        elements.append(table)

        ts = TableStyle([('GRID', (0,1), (-1,-1), 2 , colors.ReportLabBlue)       
        ])
        ts.add('BACKGROUND', (0,1), (-1,-1), colors.ReportLabLightBlue)

        table.setStyle(ts)

        doc.build(elements)

create_invoice()